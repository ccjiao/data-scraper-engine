"""
Excel 输出器 - 专业排版、号码高亮、交替行底色
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelWriter:
    """美观的Excel输出器"""

    def __init__(self, config: dict = None):
        self.config = config or {}

    def write(self, records: list[dict], output_path: str,
              columns: list = None, title: str = "",
              style_config: dict = None):
        """
        写入Excel文件

        Args:
            records: 数据记录列表
            output_path: 输出文件路径
            columns: 列定义 [{'key': 'field_name', 'label': '显示名', 'width': 15}, ...]
            title: 工作表标题
            style_config: 样式配置
        """
        if not records:
            print("  [!] 无数据可写入")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31] if title else "数据"

        # 确定列定义
        if not columns:
            columns = self._auto_columns(records)

        # 样式
        styles = self._build_styles(style_config)

        # 写表头
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def.get("label", col_def.get("key", "")))
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["header_align"]
            cell.border = styles["cell_border"]

            # 列宽
            width = col_def.get("width", 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 写数据
        for row_idx, record in enumerate(records, 2):
            for col_idx, col_def in enumerate(columns, 1):
                key = col_def.get("key", "")
                value = record.get(key, "")

                # 格式化值
                formatter = col_def.get("formatter")
                if formatter and callable(formatter):
                    value = formatter(value, record)

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = styles["cell_border"]
                cell.alignment = styles["data_align"]

                # 交替行底色
                if row_idx % 2 == 0:
                    cell.fill = styles["even_fill"]

                # 特殊样式
                style_type = col_def.get("style")
                if style_type == "front":
                    cell.font = styles["front_font"]
                elif style_type == "back":
                    cell.font = styles["back_font"]
                elif style_type == "number":
                    cell.font = styles["number_font"]

        # 冻结首行
        ws.freeze_panes = "A2"

        # 确保目录存在
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        wb.save(output_path)
        print(f"  [✓] 已保存: {output_path} ({len(records)}行 × {len(columns)}列)")

    def _auto_columns(self, records: list[dict]) -> list[dict]:
        """自动从数据推断列定义"""
        if not records:
            return []
        columns = []
        for key in records[0].keys():
            columns.append({"key": key, "label": key, "width": max(len(str(key)) * 2, 12)})
        return columns

    def _build_styles(self, style_config: dict = None) -> dict:
        """构建样式集合"""
        cfg = style_config or self.config.get("style", {})

        header_color = cfg.get("header_color", "1A1A2E")
        front_color = cfg.get("front_color", "C0392B")
        back_color = cfg.get("back_color", "2471A3")
        number_color = cfg.get("number_color", "7D3C98")
        even_color = cfg.get("even_row_color", "F7F9FC")

        thin_border = Border(
            left=Side(style="thin", color="E0E0E0"),
            right=Side(style="thin", color="E0E0E0"),
            top=Side(style="thin", color="E0E0E0"),
            bottom=Side(style="thin", color="E0E0E0"),
        )

        return {
            "header_font": Font(name="PingFang SC", size=11, bold=True, color="FFFFFF"),
            "header_fill": PatternFill(start_color=header_color, end_color=header_color, fill_type="solid"),
            "header_align": Alignment(horizontal="center", vertical="center"),
            "data_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "cell_border": thin_border,
            "even_fill": PatternFill(start_color=even_color, end_color=even_color, fill_type="solid"),
            "front_font": Font(name="PingFang SC", size=11, bold=True, color=front_color),
            "back_font": Font(name="PingFang SC", size=11, bold=True, color=back_color),
            "number_font": Font(name="PingFang SC", size=11, bold=True, color=number_color),
        }
